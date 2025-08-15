# Django core
from django.shortcuts import render
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
# Built-in and standard libraries
from decimal import Decimal
# Third-party
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.cell.cell import MergedCell

import xlsxwriter
import openpyxl
from openpyxl.utils import get_column_letter
import plotly.graph_objs as go
import plotly.offline as opy

# Project-specific
from compensation_payroll.models import RegularPayroll, SeverancePay
from compensation_payroll.services.combined.monthly_context import get_combined_monthly_detail

#
@login_required
def monthly_combined_detail(request):
    context = get_combined_monthly_detail(request)
    return render(request, 'combined_payroll/monthly_detail.html', context)


@login_required
def monthly_combined_summary(request):
    context = get_combined_monthly_detail(request)
    return render(request, 'combined_payroll/monthly_summary.html', context)


@login_required
def monthly_combined_adjustment_journal(request):
    context = get_combined_monthly_detail(request)
    return render(request, 'combined_payroll/monthly_adjustment_journal.html', context)

@login_required
def monthly_combined_adjustment_summary(request):
    context = get_combined_monthly_detail(request)
    return render(request, 'combined_payroll/monthly_adjustment_summary.html', context)


# @login_required
# def export_combined_monthly_detail_to_excel(request):
#     context = get_combined_monthly_detail(request)
#     monthly_list = context['page_obj'].object_list
#
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = "Combined Monthly Payroll"
#
#     # Styles
#     header_font = Font(bold=True, color="FFFFFF")
#     header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
#     border = Border(left=Side(style='thin'), right=Side(style='thin'),
#                     top=Side(style='thin'), bottom=Side(style='thin'))
#     money_format = '#,##0.00'
#     center_align = Alignment(horizontal="center")
#
#     row_num = 1
#
#     for item in monthly_list:
#         # Title
#         ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=7)
#         header_cell = ws.cell(row=row_num, column=1, value=f"Combined Payroll Summary for {item['month']}")
#         header_cell.font = Font(bold=True, size=14)
#         header_cell.alignment = Alignment(horizontal="center")
#         row_num += 2
#
#         # --- Regular Payroll ---
#         if any(amount != 0 for amount in item.get('regular_item_by_component', {}).values()):
#             ws.cell(row=row_num, column=1, value="Regular Payroll").font = Font(bold=True, color="0070C0")
#             row_num += 1
#
#             headers = ["Component", "Amount"]
#             for col_num, header in enumerate(headers, 1):
#                 cell = ws.cell(row=row_num, column=col_num, value=header)
#                 cell.font = header_font
#                 cell.fill = header_fill
#                 cell.alignment = center_align
#                 cell.border = border
#             row_num += 1
#
#             for comp, amount in item['regular_item_by_component'].items():
#                 if amount != 0:
#                     ws.cell(row=row_num, column=1, value=comp)
#                     amt_cell = ws.cell(row=row_num, column=2, value=float(amount))
#                     amt_cell.number_format = money_format
#                     row_num += 1
#             row_num += 1
#
#         # --- Earning Adjustments ---
#         if item.get('show_earning') and any(
#                 vals['earning_amount'] != 0 for vals in item.get('earning_adj_by_component', {}).values()):
#             ws.cell(row=row_num, column=1, value="Earning Adjustments").font = Font(bold=True, color="00B050")
#             row_num += 1
#
#             earning_headers = [
#                 "Component", "Total", "Taxable", "Non-Taxable",
#                 "Employee Pension", "Employer Pension", "Total Pension Contribution"
#             ]
#             for col_num, header in enumerate(earning_headers, 1):
#                 cell = ws.cell(row=row_num, column=col_num, value=header)
#                 cell.font = header_font
#                 cell.fill = header_fill
#                 cell.alignment = center_align
#                 cell.border = border
#             row_num += 1
#
#             for comp, vals in item['earning_adj_by_component'].items():
#                 if vals['earning_amount'] != 0:
#                     ws.cell(row=row_num, column=1, value=comp)
#                     ws.cell(row=row_num, column=2, value=float(vals['earning_amount'])).number_format = money_format
#                     ws.cell(row=row_num, column=3, value=float(vals['taxable'])).number_format = money_format
#                     ws.cell(row=row_num, column=4, value=float(vals['non_taxable'])).number_format = money_format
#                     ws.cell(row=row_num, column=5,
#                             value=float(vals['employee_pension_contribution'])).number_format = money_format
#                     ws.cell(row=row_num, column=6,
#                             value=float(vals['employer_pension_contribution'])).number_format = money_format
#                     ws.cell(row=row_num, column=7, value=float(vals['total_pension'])).number_format = money_format
#                     row_num += 1
#             row_num += 1
#
#             # Adjustment Income Tax
#             adjustment = item.get('adjustment', {})
#             if adjustment.get('employment_income_tax', 0) != 0:
#                 ws.cell(row=row_num, column=1, value="Adjustment Income Tax Summary").font = Font(bold=True,
#                                                                                                   color="7030A0")
#                 row_num += 1
#                 ws.cell(row=row_num, column=1, value="Employment Income Tax")
#                 val_cell = ws.cell(row=row_num, column=2, value=float(adjustment['employment_income_tax']))
#                 val_cell.number_format = money_format
#                 row_num += 2
#
#         # --- Deduction Adjustments ---
#         if item.get('show_deduction') and any(
#                 amount != 0 for amount in item.get('deduction_adj_by_component', {}).values()):
#             ws.cell(row=row_num, column=1, value="Deduction Adjustments").font = Font(bold=True, color="FF0000")
#             row_num += 1
#
#             deduction_headers = ["Component", "Amount"]
#             for col_num, header in enumerate(deduction_headers, 1):
#                 cell = ws.cell(row=row_num, column=col_num, value=header)
#                 cell.font = header_font
#                 cell.fill = header_fill
#                 cell.alignment = center_align
#                 cell.border = border
#             row_num += 1
#
#             for comp, amount in item['deduction_adj_by_component'].items():
#                 if amount != 0:
#                     ws.cell(row=row_num, column=1, value=comp)
#                     amt_cell = ws.cell(row=row_num, column=2, value=float(amount))
#                     amt_cell.number_format = money_format
#                     row_num += 1
#             row_num += 1
#
#         # --- Severance Payroll ---
#         severance_data = item.get('severance', {})
#         if any(val != 0 for val in severance_data.values()):
#             ws.cell(row=row_num, column=1, value="Severance Payroll").font = Font(bold=True, color="C65911")
#             row_num += 1
#
#             severance_headers = ["Component", "Amount"]
#             for col_num, header in enumerate(severance_headers, 1):
#                 cell = ws.cell(row=row_num, column=col_num, value=header)
#                 cell.font = header_font
#                 cell.fill = header_fill
#                 cell.alignment = center_align
#                 cell.border = border
#             row_num += 1
#
#             severance_items = [
#                 ("Severance Gross (Taxable)", severance_data.get('taxable_gross', 0)),
#                 ("Severance Gross", severance_data.get('gross', 0)),
#                 ("Severance Income Tax", severance_data.get('employment_income_tax', 0)),
#                 ("Total Severance Deductions", severance_data.get('total_severance_deduction', 0)),
#                 ("Severance Net Pay", severance_data.get('net', 0)),
#                 ("Severance Expense", severance_data.get('expense', 0)),
#             ]
#             for comp, amount in severance_items:
#                 if amount != 0:
#                     ws.cell(row=row_num, column=1, value=comp)
#                     amt_cell = ws.cell(row=row_num, column=2, value=float(amount))
#                     amt_cell.number_format = money_format
#                     row_num += 1
#             row_num += 1
#
#         # --- Total Summary ---
#         if any(val != 0 for val in item.get('totals', {}).values()):
#             ws.cell(row=row_num, column=1, value="Total Summary").font = Font(bold=True)
#             row_num += 1
#
#             summary_headers = ["Component", "Amount"]
#             for col_num, header in enumerate(summary_headers, 1):
#                 cell = ws.cell(row=row_num, column=col_num, value=header)
#                 cell.font = header_font
#                 cell.fill = header_fill
#                 cell.alignment = center_align
#                 cell.border = border
#             row_num += 1
#
#             summary_items = [
#                 ("Taxable Gross Pay", item['totals'].get('taxable_gross', 0)),
#                 ("Non-Taxable Gross Pay", item['totals'].get('non_taxable_gross', 0)),
#                 ("Total Gross Pay", item['totals'].get('gross', 0)),
#                 ("Total Pensionable", item['totals'].get('pensionable', 0)),
#                 ("Employee Pension", item['totals'].get('employee_pension', 0)),
#                 ("Employer Pension", item['totals'].get('employer_pension', 0)),
#                 ("Total Pension Contribution", item['totals'].get('total_pension', 0)),
#                 ("Income Tax", item['totals'].get('employment_income_tax', 0)),
#                 ("Total Deduction", item['totals'].get('total_deduction', 0)),
#                 ("Total Expense", item['totals'].get('expense', 0)),
#                 ("Final Net Pay", item['totals'].get('final_net_pay', 0)),
#             ]
#             for comp, amount in summary_items:
#                 if amount != 0:
#                     ws.cell(row=row_num, column=1, value=comp)
#                     amt_cell = ws.cell(row=row_num, column=2, value=float(amount))
#                     amt_cell.number_format = money_format
#                     row_num += 1
#             row_num += 3
#
#     # Adjust column widths
#     column_max_widths = {}
#     for row in ws.iter_rows():
#         for cell in row:
#             if not isinstance(cell, MergedCell) and cell.value:
#                 col_idx = cell.column
#                 length = len(str(cell.value))
#                 column_max_widths[col_idx] = max(column_max_widths.get(col_idx, 0), length)
#     for col_idx, max_len in column_max_widths.items():
#         col_letter = get_column_letter(col_idx)
#         ws.column_dimensions[col_letter].width = max_len + 4
#
#     # Response
#     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = 'attachment; filename=combined_monthly_payroll.xlsx'
#     wb.save(response)
#     return response


from datetime import datetime
import io

@login_required
def export_combined_monthly_detail_to_excel(request):
    # Get data from context
    context = get_combined_monthly_detail(request)
    monthly_list = context['page_obj'].object_list

    # Create workbook in memory
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Combined Monthly Payroll"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    money_format = '#,##0.00'
    center_align = Alignment(horizontal="center")
    title_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=12)

    # Color codes for different sections
    section_colors = {
        'regular': "0070C0",  # Blue
        'adjustment': "00B050",  # Green
        'deduction': "FF0000",  # Red
        'severance': "C65911",  # Orange
        'totals': "7030A0"  # Purple
    }

    # Helper function to safely convert to float
    def safe_float(value):
        try:
            return float(value) if value not in [None, ''] else 0.0
        except (ValueError, TypeError):
            return 0.0

    row_num = 1  # Start from first row

    for item in monthly_list:
        # ================================
        # 1. Title Section
        # ================================
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=7)
        title_cell = ws.cell(row=row_num, column=1, value=f"Combined Payroll Summary for {item['month']}")
        title_cell.font = title_font
        title_cell.alignment = center_align
        row_num += 3  # Extra space after title

        # ================================
        # 2. REGULAR PAYROLL SECTION
        # ================================
        if 'regular' in item and any(safe_float(val) != 0 for val in item['regular'].values()):
            ws.cell(row=row_num, column=1, value="REGULAR PAYROLL").font = Font(bold=True, size=12,
                                                                                color=section_colors['regular'])
            row_num += 2

            # Regular Components
            if 'regular_item_by_component' in item and any(
                    safe_float(amount) != 0 for amount in item['regular_item_by_component'].values()):
                ws.cell(row=row_num, column=1, value="Regular Components").font = section_font
                row_num += 1

                headers = ["Component", "Amount"]
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                    cell.alignment = center_align
                row_num += 1

                for comp, amount in item['regular_item_by_component'].items():
                    amount_float = safe_float(amount)
                    if amount_float != 0:
                        ws.cell(row=row_num, column=1, value=comp).border = thin_border
                        amt_cell = ws.cell(row=row_num, column=2, value=amount_float)
                        amt_cell.number_format = money_format
                        amt_cell.border = thin_border
                        row_num += 1
                row_num += 1

            # Regular Summary
            ws.cell(row=row_num, column=1, value="Regular Summary").font = section_font
            row_num += 1

            headers = ["Item", "Amount"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = center_align
            row_num += 1

            summary_data = [
                ("Taxable Gross", item['regular'].get('taxable_gross')),
                ("Non-Taxable Gross", item['regular'].get('non_taxable_gross')),
                ("Gross Pay", item['regular'].get('gross')),
                ("Pensionable Amount", item['regular'].get('pensionable')),
                ("Employee Pension", item['regular'].get('employee_pension')),
                ("Employer Pension", item['regular'].get('employer_pension')),
                ("Total Pension", item['regular'].get('total_pension')),
                ("Income Tax", item['regular'].get('employment_income_tax')),
                ("Total Deductions", item['regular'].get('total_regular_deduction')),
                ("Net Pay", item['regular'].get('net_pay')),
                ("Expense", item['regular'].get('expense')),
            ]

            for label, value in summary_data:
                value_float = safe_float(value)
                if value_float != 0:
                    ws.cell(row=row_num, column=1, value=label).border = thin_border
                    amt_cell = ws.cell(row=row_num, column=2, value=value_float)
                    amt_cell.number_format = money_format
                    amt_cell.border = thin_border
                    row_num += 1
            row_num += 3

        # ================================
        # 3. EARNING ADJUSTMENTS SECTION (COMPONENTS ONLY)
        # ================================
        if 'adjustment' in item and any(safe_float(val) != 0 for val in item['adjustment'].values()):
            ws.cell(row=row_num, column=1, value="EARNING ADJUSTMENTS").font = Font(bold=True, size=12,
                                                                                    color=section_colors['adjustment'])
            row_num += 2

            if 'earning_adj_by_component' in item and any(safe_float(vals.get('earning_amount', 0)) != 0 for vals in
                                                          item['earning_adj_by_component'].values()):
                ws.cell(row=row_num, column=1, value="Adjustment Components").font = section_font
                row_num += 1

                headers = [
                    "Component", "Total", "Taxable", "Non-Taxable",
                    "Employee Pension", "Employer Pension", "Total Pension"
                ]
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                    cell.alignment = center_align
                row_num += 1

                for comp, vals in item['earning_adj_by_component'].items():
                    if safe_float(vals.get('earning_amount', 0)) != 0:
                        ws.cell(row=row_num, column=1, value=comp).border = thin_border
                        ws.cell(row=row_num, column=2,
                                value=safe_float(vals.get('earning_amount', 0))).number_format = money_format
                        ws.cell(row=row_num, column=3,
                                value=safe_float(vals.get('taxable', 0))).number_format = money_format
                        ws.cell(row=row_num, column=4,
                                value=safe_float(vals.get('non_taxable', 0))).number_format = money_format
                        ws.cell(row=row_num, column=5, value=safe_float(
                            vals.get('employee_pension_contribution', 0))).number_format = money_format
                        ws.cell(row=row_num, column=6, value=safe_float(
                            vals.get('employer_pension_contribution', 0))).number_format = money_format
                        ws.cell(row=row_num, column=7,
                                value=safe_float(vals.get('total_pension', 0))).number_format = money_format
                        for col in range(1, 8):
                            ws.cell(row=row_num, column=col).border = thin_border
                        row_num += 1
                row_num += 3

        # ================================
        # 4. DEDUCTION ADJUSTMENTS
        # ================================
        if 'deduction_adj_by_component' in item and any(
                safe_float(amt) != 0 for amt in item['deduction_adj_by_component'].values()):
            ws.cell(row=row_num, column=1, value="DEDUCTION ADJUSTMENTS").font = Font(bold=True, size=12,
                                                                                      color=section_colors['deduction'])
            row_num += 2

            ws.cell(row=row_num, column=1, value="Deduction Components").font = section_font
            row_num += 1

            headers = ["Component", "Amount"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = center_align
            row_num += 1

            for comp, amount in item['deduction_adj_by_component'].items():
                amount_float = safe_float(amount)
                if amount_float != 0:
                    ws.cell(row=row_num, column=1, value=comp).border = thin_border
                    amt_cell = ws.cell(row=row_num, column=2, value=amount_float)
                    amt_cell.number_format = money_format
                    amt_cell.border = thin_border
                    row_num += 1
            row_num += 2

        # ================================
        # 5. ADJUSTMENT SUMMARY (NOW BELOW DEDUCTION)
        # ================================
        if 'adjustment' in item and any(safe_float(val) != 0 for val in item['adjustment'].values()):
            ws.cell(row=row_num, column=1, value="ADJUSTMENT SUMMARY").font = section_font
            row_num += 1

            headers = ["Item", "Amount"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = center_align
            row_num += 1

            adj_summary = [
                ("Taxable Gross", item['adjustment'].get('taxable_gross')),
                ("Non-Taxable Gross", item['adjustment'].get('non_taxable_gross')),
                ("Gross Adjustment", item['adjustment'].get('gross')),
                ("Adjusted Pensionable", item['adjustment'].get('adjusted_pensionable')),
                ("Employee Pension", item['adjustment'].get('employee_pension')),
                ("Employer Pension", item['adjustment'].get('employer_pension')),
                ("Total Pension", item['adjustment'].get('total_pension')),
                ("Income Tax", item['adjustment'].get('employment_income_tax')),
                ("Earning Adjustment Deduction", item['adjustment'].get('earning_adjustment_deduction')),
                ("Net Monthly Adjustment", item['adjustment'].get('net_monthly_adjustment')),
                ("Expense", item['adjustment'].get('expense')),
            ]

            for label, value in adj_summary:
                value_float = safe_float(value)
                if value_float != 0:
                    ws.cell(row=row_num, column=1, value=label).border = thin_border
                    amt_cell = ws.cell(row=row_num, column=2, value=value_float)
                    amt_cell.number_format = money_format
                    amt_cell.border = thin_border
                    row_num += 1
            row_num += 3

        # ================================
        # 6. SEVERANCE PAYROLL
        # ================================
        if 'severance' in item and any(safe_float(val) != 0 for val in item['severance'].values()):
            ws.cell(row=row_num, column=1, value="SEVERANCE PAYROLL").font = Font(bold=True, size=12,
                                                                                  color=section_colors['severance'])
            row_num += 2

            ws.cell(row=row_num, column=1, value="Severance Details").font = section_font
            row_num += 1

            headers = ["Item", "Amount"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = center_align
            row_num += 1

            severance_items = [
                ("Severance Gross (Taxable)", item['severance'].get('taxable_gross')),
                ("Severance Gross", item['severance'].get('gross')),
                ("Severance Income Tax", item['severance'].get('employment_income_tax')),
                ("Total Severance Deductions", item['severance'].get('total_severance_deduction')),
                ("Severance Net Pay", item['severance'].get('net')),
                ("Severance Expense", item['severance'].get('expense')),
            ]

            for label, value in severance_items:
                value_float = safe_float(value)
                if value_float != 0:
                    ws.cell(row=row_num, column=1, value=label).border = thin_border
                    amt_cell = ws.cell(row=row_num, column=2, value=value_float)
                    amt_cell.number_format = money_format
                    amt_cell.border = thin_border
                    row_num += 1
            row_num += 3

        # ================================
        # 7. TOTALS
        # ================================
        if 'totals' in item and any(safe_float(val) != 0 for val in item['totals'].values()):
            ws.cell(row=row_num, column=1, value="TOTAL SUMMARY").font = Font(bold=True, size=12,
                                                                              color=section_colors['totals'])
            row_num += 2

            headers = ["Item", "Amount"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = center_align
            row_num += 1

            summary_items = [
                ("Taxable Gross Pay", item['totals'].get('taxable_gross')),
                ("Non-Taxable Gross Pay", item['totals'].get('non_taxable_gross')),
                ("Total Gross Pay", item['totals'].get('gross')),
                ("Total Pensionable", item['totals'].get('pensionable')),
                ("Employee Pension", item['totals'].get('employee_pension')),
                ("Employer Pension", item['totals'].get('employer_pension')),
                ("Total Pension Contribution", item['totals'].get('total_pension')),
                ("Income Tax", item['totals'].get('employment_income_tax')),
                ("Total Deduction", item['totals'].get('total_deduction')),
                ("Total Expense", item['totals'].get('expense')),
                ("Final Net Pay", item['totals'].get('final_net_pay')),
            ]

            for label, value in summary_items:
                value_float = safe_float(value)
                if value_float != 0:
                    ws.cell(row=row_num, column=1, value=label).border = thin_border
                    amt_cell = ws.cell(row=row_num, column=2, value=value_float)
                    amt_cell.number_format = money_format
                    amt_cell.border = thin_border
                    row_num += 1
            row_num += 3

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"combined_monthly_payroll_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'

    # Save workbook to response
    wb.save(response)

    return response


@login_required
def export_combined_monthly_summary_to_excel(request):
    # Get data from context
    context = get_combined_monthly_detail(request)
    monthly_list = context['page_obj'].object_list

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Payroll Summary"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
    money_format = '#,##0.00'
    center_align = Alignment(horizontal="center")
    title_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=12)

    # Section colors
    section_colors = {
        'regular': "0070C0",
        'adjustment': "00B050",
        'severance': "C65911",
        'totals': "7030A0"
    }

    def safe_float(value):
        try:
            return float(value) if value not in [None, ''] else 0.0
        except:
            return 0.0

    row_num = 1

    for item in monthly_list:
        # Title
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=3)
        title_cell = ws.cell(row=row_num, column=1, value=f"Monthly Payroll Summary for {item['month']}")
        title_cell.font = title_font
        title_cell.alignment = center_align
        row_num += 3

        # -------------------------------
        # 1. Regular Payroll
        # -------------------------------
        if 'regular' in item and any(safe_float(v) != 0 for v in item['regular'].values()):
            ws.cell(row=row_num, column=1, value="REGULAR PAYROLL").font = Font(bold=True, size=12, color=section_colors['regular'])
            row_num += 2

            headers = ["Item", "Amount"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = center_align
            row_num += 1

            summary_data = [
                ("Taxable Gross", item['regular'].get('taxable_gross')),
                ("Non-Taxable Gross", item['regular'].get('non_taxable_gross')),
                ("Gross Pay", item['regular'].get('gross')),
                ("Pensionable Amount", item['regular'].get('pensionable')),
                ("Employee Pension", item['regular'].get('employee_pension')),
                ("Employer Pension", item['regular'].get('employer_pension')),
                ("Total Pension", item['regular'].get('total_pension')),
                ("Income Tax", item['regular'].get('employment_income_tax')),
                ("Total Deductions", item['regular'].get('total_regular_deduction')),
                ("Net Pay", item['regular'].get('net_pay')),
                ("Expense", item['regular'].get('expense')),
            ]

            for label, value in summary_data:
                value_float = safe_float(value)
                if value_float != 0:
                    ws.cell(row=row_num, column=1, value=label).border = thin_border
                    amt_cell = ws.cell(row=row_num, column=2, value=value_float)
                    amt_cell.number_format = money_format
                    amt_cell.border = thin_border
                    row_num += 1
            row_num += 3

        # -------------------------------
        # 2. Earning Adjustments (Summary only)
        # -------------------------------
        if 'adjustment' in item and any(safe_float(v) != 0 for v in item['adjustment'].values()):
            ws.cell(row=row_num, column=1, value="EARNING ADJUSTMENTS").font = Font(bold=True, size=12, color=section_colors['adjustment'])
            row_num += 2

            headers = ["Item", "Amount"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = center_align
            row_num += 1

            adj_summary = [
                ("Taxable Gross", item['adjustment'].get('taxable_gross')),
                ("Non-Taxable Gross", item['adjustment'].get('non_taxable_gross')),
                ("Gross Adjustment", item['adjustment'].get('gross')),
                ("Adjusted Pensionable", item['adjustment'].get('adjusted_pensionable')),
                ("Employee Pension", item['adjustment'].get('employee_pension')),
                ("Employer Pension", item['adjustment'].get('employer_pension')),
                ("Total Pension", item['adjustment'].get('total_pension')),
                ("Income Tax", item['adjustment'].get('employment_income_tax')),
                ("Earning Adjustment Deduction", item['adjustment'].get('earning_adjustment_deduction')),
                ("Net Monthly Adjustment", item['adjustment'].get('net_monthly_adjustment')),
                ("Expense", item['adjustment'].get('expense')),
            ]

            for label, value in adj_summary:
                value_float = safe_float(value)
                if value_float != 0:
                    ws.cell(row=row_num, column=1, value=label).border = thin_border
                    amt_cell = ws.cell(row=row_num, column=2, value=value_float)
                    amt_cell.number_format = money_format
                    amt_cell.border = thin_border
                    row_num += 1
            row_num += 3

        # -------------------------------
        # 3. Severance Payroll
        # -------------------------------
        if 'severance' in item and any(safe_float(v) != 0 for v in item['severance'].values()):
            ws.cell(row=row_num, column=1, value="SEVERANCE PAYROLL").font = Font(bold=True, size=12, color=section_colors['severance'])
            row_num += 2

            headers = ["Item", "Amount"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = center_align
            row_num += 1

            severance_items = [
                ("Severance Gross (Taxable)", item['severance'].get('taxable_gross')),
                ("Severance Gross", item['severance'].get('gross')),
                ("Severance Income Tax", item['severance'].get('employment_income_tax')),
                ("Total Severance Deductions", item['severance'].get('total_severance_deduction')),
                ("Severance Net Pay", item['severance'].get('net')),
                ("Severance Expense", item['severance'].get('expense')),
            ]

            for label, value in severance_items:
                value_float = safe_float(value)
                if value_float != 0:
                    ws.cell(row=row_num, column=1, value=label).border = thin_border
                    amt_cell = ws.cell(row=row_num, column=2, value=value_float)
                    amt_cell.number_format = money_format
                    amt_cell.border = thin_border
                    row_num += 1
            row_num += 3

        # -------------------------------
        # 4. Total Summary
        # -------------------------------
        if 'totals' in item and any(safe_float(v) != 0 for v in item['totals'].values()):
            ws.cell(row=row_num, column=1, value="TOTAL SUMMARY").font = Font(bold=True, size=12, color=section_colors['totals'])
            row_num += 2

            headers = ["Item", "Amount"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = center_align
            row_num += 1

            summary_items = [
                ("Taxable Gross Pay", item['totals'].get('taxable_gross')),
                ("Non-Taxable Gross Pay", item['totals'].get('non_taxable_gross')),
                ("Total Gross Pay", item['totals'].get('gross')),
                ("Total Pensionable", item['totals'].get('pensionable')),
                ("Employee Pension", item['totals'].get('employee_pension')),
                ("Employer Pension", item['totals'].get('employer_pension')),
                ("Total Pension Contribution", item['totals'].get('total_pension')),
                ("Income Tax", item['totals'].get('employment_income_tax')),
                ("Total Deduction", item['totals'].get('total_deduction')),
                ("Total Expense", item['totals'].get('expense')),
                ("Final Net Pay", item['totals'].get('final_net_pay')),
            ]

            for label, value in summary_items:
                value_float = safe_float(value)
                if value_float != 0:
                    ws.cell(row=row_num, column=1, value=label).border = thin_border
                    amt_cell = ws.cell(row=row_num, column=2, value=value_float)
                    amt_cell.number_format = money_format
                    amt_cell.border = thin_border
                    row_num += 1
            row_num += 3

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = (max_length + 2) * 1.2

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"combined_monthly_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'

    wb.save(response)
    return response



#graph combined monthly
@login_required
def combined_monthly_graph_view(request):
    monthly_summary = get_combined_monthly_detail(request).get("monthly_summary", {})

    # Order by "Month-YYYY"
    def parse_key(key):
        try:
            month_str, year_str = key.split('-')
            year = int(year_str)
            month_order_full = {
                "January": 1, "February": 2, "March": 3, "April": 4,
                "May": 5, "June": 6, "July": 7, "August": 8,
                "September": 9, "October": 10, "November": 11, "December": 12
            }
            month_num = month_order_full.get(month_str.capitalize(), 0)
            return (year, month_num)
        except Exception:
            return (0, 0)

    sorted_summary = sorted(monthly_summary.items(), key=lambda x: parse_key(x[0]))

    # Components to include in the bar chart
    components = [
        'taxable_gross',
        'non_taxable_gross',
        'gross',
        'pensionable',
        'employee_pension',
        'employer_pension',
        'total_pension',
        'employment_income_tax',
        'expense',
        'final_net_pay',
    ]

    months = []
    data_by_component = {comp: [] for comp in components}

    for month_key, data in sorted_summary:
        months.append(month_key)  # Use raw key like "July-2024"
        totals = data['totals']
        for comp in components:
            data_by_component[comp].append(float(totals.get(comp, 0)))

    # Create Bar chart traces
    bar_traces = [
        go.Bar(x=months, y=data_by_component[comp], name=comp.replace('_', ' ').title())
        for comp in components
    ]

    bar_layout = go.Layout(
        title='Monthly Payroll Summary By Components',
        barmode='group',
        xaxis=dict(title='Month'),
        yaxis=dict(title='Amount (ETB)'),
        template='plotly_white',
        margin=dict(l=40, r=40, t=60, b=60),
    )

    bar_fig = go.Figure(data=bar_traces, layout=bar_layout)
    bar_div = opy.plot(bar_fig, auto_open=False, output_type='div')

    # --- Pie chart for selected month ---
    selected_month = request.GET.get('month')
    if not selected_month and sorted_summary:
        selected_month = sorted_summary[-1][0]

    month_options = [key for key, _ in sorted_summary]
    selected_data = dict(sorted_summary).get(selected_month)

    if selected_data:
        totals = selected_data['totals']
        gross = float(totals.get('gross', 0))
        net = float(totals.get('final_net_pay', 0))
        emp_pension = float(totals.get('employer_pension', 0))
        emp_deduction = gross - net

        labels = [
            'Net Pay (Employee Take-home)',
            'Employee Deductions (Tax, Pension, etc.)',
            'Employer Pension Contribution',
        ]
        values = [net, emp_deduction, emp_pension]

        if gross <= 0 or sum(values) <= 0:
            pie_div = "<p>Insufficient data for pie chart.</p>"
        else:
            pie_trace = go.Pie(
                labels=labels,
                values=values,
                hole=0.3,
                hoverinfo='label+percent+value',
                textinfo='label+percent'
            )
            pie_layout = go.Layout(
                title=f'Payroll Cost Breakdown for {selected_month}',
                template='plotly_white',
                margin=dict(l=40, r=40, t=60, b=60),
            )
            pie_fig = go.Figure(data=[pie_trace], layout=pie_layout)
            pie_div = opy.plot(pie_fig, auto_open=False, output_type='div')
    else:
        pie_div = "<p>No data available for selected month.</p>"

    return render(request, 'combined_payroll/monthly_graph.html', {
        'month_options': month_options,
        'selected_month': selected_month,
        'pie_div': pie_div,
        'bar_div': bar_div,
    })