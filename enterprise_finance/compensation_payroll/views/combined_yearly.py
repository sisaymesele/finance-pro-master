from django.shortcuts import render
from django.http import HttpResponse
from decimal import Decimal
from django.contrib.auth.decorators import login_required
from compensation_payroll.services.combined.yearly_context import get_combined_yearly_detail
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from openpyxl import Workbook

@login_required
def yearly_combined_detail_view(request):
    context = get_combined_yearly_detail(request)
    return render(request, 'combined_payroll/yearly_detail.html', context)

#monthly summary
@login_required
def yearly_combined_summary_view(request):

    context = get_combined_yearly_detail(request)

    return render(request, 'combined_payroll/yearly_summary.html', context)

#
@login_required
def export_combined_yearly_detail_to_excel(request):
    context = get_combined_yearly_detail(request)
    yearly_list = context['page_obj'].object_list

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Combined Yearly Payroll"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    money_format = '#,##0.00'
    center_align = Alignment(horizontal="center")

    row_num = 1

    for item in yearly_list:
        # Skip year entirely if no section has data
        if not (
                any(v != 0 for v in item.get('regular_item_by_component', {}).values()) or
                any(v['earning_amount'] != 0 for v in item.get('earning_adj_by_component', {}).values()) or
                item.get('adjustment', {}).get('employment_income_tax', 0) != 0 or
                any(v != 0 for v in item.get('deduction_adj_by_component', {}).values()) or
                any(v != 0 for v in item.get('severance', {}).values()) or
                any(v != 0 for v in item.get('totals', {}).values())
        ):
            continue

        # Title
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=7)
        header_cell = ws.cell(row=row_num, column=1, value=f"Combined Payroll Summary for {item['year']}")
        header_cell.font = Font(bold=True, size=14)
        header_cell.alignment = Alignment(horizontal="center")
        row_num += 2

        # --- Regular Payroll ---
        if any(amount != 0 for amount in item.get('regular_item_by_component', {}).values()):
            ws.cell(row=row_num, column=1, value="Regular Payroll").font = Font(bold=True, color="0070C0")
            row_num += 1

            headers = ["Component", "Amount"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = border
            row_num += 1

            for comp, amount in item['regular_item_by_component'].items():
                if amount != 0:
                    ws.cell(row=row_num, column=1, value=comp)
                    amt_cell = ws.cell(row=row_num, column=2, value=float(amount))
                    amt_cell.number_format = money_format
                    row_num += 1
            row_num += 1

        # --- Earning Adjustments ---
        if item.get('show_earning') and any(
                vals['earning_amount'] != 0 for vals in item.get('earning_adj_by_component', {}).values()):
            ws.cell(row=row_num, column=1, value="Earning Adjustments").font = Font(bold=True, color="00B050")
            row_num += 1

            earning_headers = [
                "Component", "Total", "Taxable", "Non-Taxable",
                "Employee Pension", "Employer Pension", "Total Pension Contribution"
            ]
            for col_num, header in enumerate(earning_headers, 1):
                cell = ws.cell(row=row_num, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = border
            row_num += 1

            for comp, vals in item['earning_adj_by_component'].items():
                if vals['earning_amount'] != 0:
                    ws.cell(row=row_num, column=1, value=comp)
                    ws.cell(row=row_num, column=2, value=float(vals['earning_amount'])).number_format = money_format
                    ws.cell(row=row_num, column=3, value=float(vals['taxable'])).number_format = money_format
                    ws.cell(row=row_num, column=4, value=float(vals['non_taxable'])).number_format = money_format
                    ws.cell(row=row_num, column=5,
                            value=float(vals['employee_pension_contribution'])).number_format = money_format
                    ws.cell(row=row_num, column=6,
                            value=float(vals['employer_pension_contribution'])).number_format = money_format
                    ws.cell(row=row_num, column=7, value=float(vals['total_pension'])).number_format = money_format
                    row_num += 1
            row_num += 1

            # Adjustment Income Tax
            adjustment = item.get('adjustment', {})
            if adjustment.get('employment_income_tax', 0) != 0:
                ws.cell(row=row_num, column=1, value="Adjustment Income Tax Summary").font = Font(bold=True,
                                                                                                  color="7030A0")
                row_num += 1
                ws.cell(row=row_num, column=1, value="Employment Income Tax")
                val_cell = ws.cell(row=row_num, column=2, value=float(adjustment['employment_income_tax']))
                val_cell.number_format = money_format
                row_num += 2

        # --- Deduction Adjustments ---
        if item.get('show_deduction') and any(
                amount != 0 for amount in item.get('deduction_adj_by_component', {}).values()):
            ws.cell(row=row_num, column=1, value="Deduction Adjustments").font = Font(bold=True, color="FF0000")
            row_num += 1

            deduction_headers = ["Component", "Amount"]
            for col_num, header in enumerate(deduction_headers, 1):
                cell = ws.cell(row=row_num, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = border
            row_num += 1

            for comp, amount in item['deduction_adj_by_component'].items():
                if amount != 0:
                    ws.cell(row=row_num, column=1, value=comp)
                    amt_cell = ws.cell(row=row_num, column=2, value=float(amount))
                    amt_cell.number_format = money_format
                    row_num += 1
            row_num += 1

        # --- Severance Payroll ---
        severance_data = item.get('severance', {})
        if any(val != 0 for val in severance_data.values()):
            ws.cell(row=row_num, column=1, value="Severance Payroll").font = Font(bold=True, color="C65911")
            row_num += 1

            severance_headers = ["Component", "Amount"]
            for col_num, header in enumerate(severance_headers, 1):
                cell = ws.cell(row=row_num, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = border
            row_num += 1

            severance_items = [
                ("Severance Gross (Taxable)", severance_data.get('taxable_gross', 0)),
                ("Severance Gross", severance_data.get('gross', 0)),
                ("Severance Income Tax", severance_data.get('employment_income_tax', 0)),
                ("Total Severance Deductions", severance_data.get('total_severance_deduction', 0)),
                ("Severance Net Pay", severance_data.get('net', 0)),
                ("Severance Expense", severance_data.get('expense', 0)),
            ]
            for comp, amount in severance_items:
                if amount != 0:
                    ws.cell(row=row_num, column=1, value=comp)
                    amt_cell = ws.cell(row=row_num, column=2, value=float(amount))
                    amt_cell.number_format = money_format
                    row_num += 1
            row_num += 1

        # --- Total Summary ---
        if any(val != 0 for val in item.get('totals', {}).values()):
            ws.cell(row=row_num, column=1, value="Total Summary").font = Font(bold=True)
            row_num += 1

            summary_headers = ["Component", "Amount"]
            for col_num, header in enumerate(summary_headers, 1):
                cell = ws.cell(row=row_num, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = border
            row_num += 1

            summary_items = [
                ("Taxable Gross Pay", item['totals'].get('taxable_gross', 0)),
                ("Non-Taxable Gross Pay", item['totals'].get('non_taxable_gross', 0)),
                ("Total Gross Pay", item['totals'].get('gross', 0)),
                ("Total Pensionable", item['totals'].get('pensionable', 0)),
                ("Employee Pension", item['totals'].get('employee_pension', 0)),
                ("Employer Pension", item['totals'].get('employer_pension', 0)),
                ("Total Pension Contribution", item['totals'].get('total_pension', 0)),
                ("Income Tax", item['totals'].get('employment_income_tax', 0)),
                ("Total Deduction", item['totals'].get('total_deduction', 0)),
                ("Total Expense", item['totals'].get('expense', 0)),
                ("Final Net Pay", item['totals'].get('final_net_pay', 0)),
            ]
            for comp, amount in summary_items:
                if amount != 0:
                    ws.cell(row=row_num, column=1, value=comp)
                    amt_cell = ws.cell(row=row_num, column=2, value=float(amount))
                    amt_cell.number_format = money_format
                    row_num += 1
            row_num += 3

    # Adjust column widths
    column_max_widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if not isinstance(cell, MergedCell) and cell.value:
                col_idx = cell.column
                length = len(str(cell.value))
                column_max_widths[col_idx] = max(column_max_widths.get(col_idx, 0), length)
    for col_idx, max_len in column_max_widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max_len + 4

    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=combined_yearly_payroll.xlsx'
    wb.save(response)
    return response
    

#yearly detail excel
@login_required
def export_combined_yearly_summary_to_excel(request):
    context = get_combined_yearly_detail(request)
    yearly_summary = context.get('yearly_summary', {})

    section_fields = {
        'regular': {
            'taxable_gross': 'Taxable Gross',
            'non_taxable_gross': 'Non-Taxable Gross',
            'gross': 'Gross Pay',
            'pensionable': 'Pensionable',
            'employee_pension': 'Employee Pension',
            'employer_pension': 'Employer Pension',
            'total_pension': 'Total Pension Contribution',
            'employment_income_tax': 'Income Tax',
            'total_regular_deduction': 'Total Deductions',
            'net_pay': 'Net Pay',
            'expense': 'Expense',
        },
        'adjustment': {
            'taxable_gross': 'Adjusted Taxable Gross',
            'non_taxable_gross': 'Adjusted Non-Taxable Gross',
            'gross': 'Adjusted Gross Pay',
            'adjusted_pensionable': 'Adjusted Pensionable',
            'employee_pension': 'Adjusted Employee Pension',
            'employer_pension': 'Adjusted Employer Pension',
            'total_pension': 'Adjusted Total Pension Contribution',
            'employment_income_tax': 'Income Tax on Adjustment',
            'total_adjustment_deduction': 'Adjustment Deductions',
            'expense': 'Adjusted Expense',
        },
        'severance': {
            'taxable_gross': 'Severance Gross (Taxable)',
            'gross': 'Severance Gross',
            'employment_income_tax': 'Severance Income Tax',
            'total_severance_deduction': 'Total Severance Deductions',
            'net': 'Severance Net Pay',
            'expense': 'Severance Expense',
        },
        'totals': {
            'taxable_gross': 'Total Taxable Gross',
            'non_taxable_gross': 'Total Non-Taxable Gross',
            'gross': 'Total Gross Pay',
            'pensionable': 'Total Pensionable',
            'employee_pension': 'Total Employee Pension',
            'employer_pension': 'Total Employer Pension',
            'total_pension': 'Total Pension Contribution',
            'employment_income_tax': 'Total Income Tax',
            'total_deduction': 'Total Deductions',
            'expense': 'Total Expense',
            'final_net_pay': 'Final Net Pay',
        }
    }

    section_colors = {
        'regular': 'BDD7EE',
        'adjustment': 'FDE9D9',
        'severance': 'F8CBAD',
        'totals': 'D9EAD3',
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "Yearly Payroll Summary"
    row = 1

    for year, data in yearly_summary.items():
        year_cell = ws.cell(row=row, column=1, value=f"{year} Payroll Summary")
        year_cell.font = Font(bold=True, size=14)
        row += 2

        for section_title, key in [('Regular', 'regular'), ('Adjustment', 'adjustment'),
                                   ('Severance', 'severance'), ('Totals', 'totals')]:
            section_data = data.get(key, {})
            labels = section_fields.get(key, {})

            # Skip section if all values are None or 0
            non_zero = any(
                v is not None and float(v) != 0.0
                for k, v in section_data.items()
                if k in labels
            )
            if not non_zero:
                continue

            # Section Header
            sec_cell = ws.cell(row=row, column=1, value=section_title)
            sec_cell.font = Font(bold=True, size=12)
            fill_color = section_colors.get(key)
            if fill_color:
                sec_cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            row += 1

            # Table Header
            comp_cell = ws.cell(row=row, column=1, value="Component")
            amt_cell = ws.cell(row=row, column=2, value="Amount")
            comp_cell.font = amt_cell.font = Font(bold=True)
            comp_cell.fill = amt_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            comp_cell.alignment = amt_cell.alignment = Alignment(horizontal="center")
            row += 1

            # Data rows with conditional check
            for field_key, label in labels.items():
                value = section_data.get(field_key)
                if value is None or float(value) == 0.0:
                    continue

                amount = float(value)
                ws.cell(row=row, column=1, value=label)
                amt_cell = ws.cell(row=row, column=2, value=amount)
                amt_cell.number_format = '#,##0.00'
                row += 1

            row += 3  # Extra space between sections

        row += 2  # Space between years

    # Adjust column widths
    for col in range(1, 3):
        max_length = 0
        col_letter = get_column_letter(col)
        for cell in ws[col_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 5

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Yearly_Payroll_Summary.xlsx"'
    wb.save(response)
    return response

