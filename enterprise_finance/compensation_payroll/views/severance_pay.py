from django.shortcuts import render, get_object_or_404, redirect
from django.core.paginator import Paginator
from django.urls import reverse
from django.contrib import messages
from django.db.models import Sum, F, Q
from datetime import datetime
from django.contrib.auth.decorators import login_required
from compensation_payroll.models import SeverancePay
from compensation_payroll.forms import SeverancePayForm

# export
import openpyxl
from io import BytesIO
from django.http import HttpResponse



# start severance
@login_required
def severance_pay_list(request):
    severance_pays = SeverancePay.objects.filter(organization_name=request.user.organization_name)
    # Search functionality (if needed in the future)
    search_query = request.GET.get('search', '').strip()
    if search_query:
        severance_pays = severance_pays.filter(
            Q(year__icontains=search_query) | Q(month__icontains=search_query) |
            Q(personnel_full_name__personnel_id__icontains=search_query) | Q(personnel_full_name__first_name__icontains=search_query) |
            Q(personnel_full_name__father_name__icontains=search_query) | Q(personnel_full_name__last_name__icontains=search_query)
        )
    severance_pays = severance_pays.order_by('-id')

    # Pagination
    paginator = Paginator(severance_pays, 10)  # Show 4 items per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Pass data to the template
    context = {'page_obj': page_obj, }
    return render(request, 'severance_pay/list.html', context)


@login_required
def severance_pay_detail(request, pk):
    # Fetch the specific payroll record by its ID
    severance_pay = get_object_or_404(SeverancePay, pk=pk, organization_name=request.user.organization_name)

    # Context data to pass to the template
    context = {
        'severance_pay': severance_pay,  # The detailed payroll record
    }

    # Render the template with the context data
    return render(request, 'severance_pay/detail.html', context)


@login_required
def create_severance_pay(request):
    if request.method == 'POST':
        form = SeverancePayForm(request.POST, request=request)
        if form.is_valid():
            severance_pay = form.save(commit=False)
            severance_pay.organization_name = request.user.organization_name
            severance_pay.save()
            messages.success(request, "Severance pay created successfully!")
            return redirect('severance-pay-list')
        else:
            messages.error(request, "Error creating severance pay. Check the form.")
    else:
        form = SeverancePayForm(request=request)

    context = {
        'form': form,
        'form_title': 'Create Severance Pay',
        'submit_button_text': 'Create Severance Pay',
        'back_url': reverse('severance-pay-list'),
        'show_add_personnel': True,  # Enable Add Personnel button
        'current_url': request.path  # Pass current URL for redirect back
    }
    return render(request, 'severance_pay/form.html', context)


@login_required
def update_severance_pay(request, pk):
    # Fetch the severance pay to update, ensure it belongs to the logged-in user
    severance_pay = get_object_or_404(SeverancePay, pk=pk, organization_name=request.user.organization_name)

    if request.method == 'POST':
        form = SeverancePayForm(request.POST, instance=severance_pay, request=request)
        if form.is_valid():
            form.save()
            messages.success(request, "Severance pay updated successfully!")
            return redirect('severance-pay-list')
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = SeverancePayForm(instance=severance_pay, request=request)

    context = {
        'form': form,
        'form_title': 'Update Severance Pay',
        'submit_button_text': 'Update Severance Pay',
        'back_url': reverse('severance-pay-list'),
        'show_add_personnel': True,  # Enable Add Personnel button
        'current_url': request.path  # Pass current URL for redirect back
    }
    return render(request, 'severance_pay/form.html', context)


@login_required
def delete_severance_pay(request, pk):
    # Fetch the severance pay to delete, ensure it belongs to the logged-in user
    severance_pay = get_object_or_404(SeverancePay, pk=pk, organization_name=request.user.organization_name)

    # If the form is submitted via POST, delete the record
    if request.method == "POST":
        severance_pay.delete()
        messages.success(request, "Severance pay deleted successfully!")
        return redirect('severance-pay-list')  # Redirect to severance pay list page

    # If not POST, render a confirmation page
    context = {'severance_pay': severance_pay}
    return render(request, 'severance_pay/delete_confirm.html', context)  # Render confirmation page


# export
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

@login_required
def export_severance_pay_to_excel(request):
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Severance Pay"

    # Exclude unnecessary fields
    excluded_fields = {'id'}
    field_names = [
        field.name for field in SeverancePay._meta.get_fields()
        if not field.is_relation and field.name not in excluded_fields
    ]

    # Convert field names to "Title Case With Line Breaks"
    def format_header(name):
        parts = name.replace("_", " ").title().split()
        # Join with line breaks every max 3 words
        lines = [' '.join(parts[i:i+3]) for i in range(0, len(parts), 3)]
        return "\n".join(lines)

    header_titles = [format_header(name) for name in field_names]
    ws.append(header_titles)

    # Style headers
    header_fill = PatternFill(start_color="FF0070C0", end_color="FF0070C0", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Fetch and write data rows
    severance_pays = SeverancePay.objects.filter(
        organization_name=request.user.organization_name
    ).values_list(*field_names)

    for row_data in severance_pays:
        cleaned_row = [
            value.replace(tzinfo=None) if isinstance(value, datetime) else value
            for value in row_data
        ]
        ws.append(cleaned_row)

    # Auto-fit column widths
    for col_idx, column_cells in enumerate(ws.columns, 1):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        adjusted_width = min(max(max_length + 2, 12), 15)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # Prepare HTTP response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=severance_payrolls.xlsx"

    # Save workbook to response
    wb.save(response)
    return response


# aggregate

@login_required
def severance_pay_report(request, template_name):
    # Filter by logged-in user
    user_severance_pay = SeverancePay.objects.filter(organization_name=request.user.organization_name)

    # Aggregated by Year-Month
    severance_summary_month_year = user_severance_pay.values('year', 'month').annotate(
        total_gross_severance=Sum('gross_severance_pay'),
        total_employment_income_tax_from_severance_pay=Sum('employment_income_tax_from_severance_pay'),
        total_net_severance_pay=Sum('net_severance_pay')
    ).order_by('year', 'month')

    # Aggregated by Year
    severance_summary_year = user_severance_pay.values('year').annotate(
        total_gross_severance=Sum('gross_severance_pay'),
        total_employment_income_tax_from_severance_pay=Sum('employment_income_tax_from_severance_pay'),
        total_net_severance_pay=Sum('net_severance_pay')
    ).order_by('year')

    return render(request, template_name, {
        'severance_summary_month_year': severance_summary_month_year,
        'severance_summary_year': severance_summary_year
    })
