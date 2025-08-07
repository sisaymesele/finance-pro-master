from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
#exel
from django.http import HttpResponse
from openpyxl import Workbook
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from compensation_payroll.models import DeductionAdjustment
from compensation_payroll.forms import DeductionAdjustmentForm
from compensation_payroll.services.deduction_adjustment.context import get_deduction_adjustment_context
from compensation_payroll.services.excel_export import ExportUtilityService


@login_required
def deduction_object_list(request):
    context = get_deduction_adjustment_context(request)
    return render(request, 'deduction_adjustment/list.html', context)


@login_required
def deduction_object_detail(request):
    context = get_deduction_adjustment_context(request)
    return render(request, 'deduction_adjustment/detail.html', context)


@login_required
def deduction_per_adjusted_month(request):
    context = get_deduction_adjustment_context(request)
    return render(request, 'deduction_adjustment/per_adjusted_month.html', context)


@login_required
def monthly_deduction_adjustment(request):
    context = get_deduction_adjustment_context(request)
    return render(request, 'deduction_adjustment/monthly_deduction.html', context)



# Create View
@login_required
def create_deduction_adjustment(request):
    if request.method == 'POST':
        form = DeductionAdjustmentForm(request.POST, request=request)
        if form.is_valid():
            deduction_adjustment = form.save(commit=False)
            deduction_adjustment.organization_name = request.user.organization_name
            deduction_adjustment.save()
            messages.success(request, "Deduction Adjustment created successfully!")
            return redirect('deduction_adjustment_list')
        else:
            messages.error(request, "Error creating the deduction adjustment. Check the form.")
    else:
        form = DeductionAdjustmentForm(request=request)

    context = {
        'form': form,
        'form_title': 'Create Deduction Adjustment',
        'submit_button_text': 'Create Deduction Adjustment',
    }
    return render(request, 'deduction_adjustment/form.html', context)


# Update View
@login_required
def update_deduction_adjustment(request, pk):
    deduction_adjustment = get_object_or_404(DeductionAdjustment, pk=pk,
                                             organization_name=request.user.organization_name)

    if request.method == 'POST':
        form = DeductionAdjustmentForm(request.POST, instance=deduction_adjustment, request=request)
        if form.is_valid():
            form.save()
            messages.success(request, "Deduction Adjustment updated successfully!")
            return redirect('deduction_adjustment_list')
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = DeductionAdjustmentForm(instance=deduction_adjustment, request=request)

    context = {
        'form': form,
        'form_title': 'Update Deduction Adjustment',
        'submit_button_text': 'Update Deduction Adjustment',
    }
    return render(request, 'deduction_adjustment/form.html', context)


# Delete View
@login_required
def delete_deduction_adjustment(request, pk):
    deduction_adjustment = get_object_or_404(DeductionAdjustment, pk=pk,
                                             organization_name=request.user.organization_name)

    if request.method == "POST":
        deduction_adjustment.delete()
        messages.success(request, "Deduction Adjustment deleted successfully!")
        return redirect('deduction_adjustment_detail')

    context = {'deduction_adjustment': deduction_adjustment}

    return render(request, 'deduction_adjustment/delete_confirm.html', context)


# export deduction adjustment
@login_required
def export_deduction_adjustment_list_to_excel(request):
    context = get_deduction_adjustment_context(request)
    deductions = context.get('deduction_adjustments', [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Deduction Adjustments"

    headers = [
        "#", "Record Month", "Adjusted Payroll Month",
        "First Name", "Father Name", "Last Name",
        "Case", "Component", "Deduction Amount",
        "Period Start", "Period End", "Months Covered",
        "Created At", "Updated At"
    ]

    total_columns = len(headers)

    # Title row (1st)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "Deduction Adjustment Report"
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Subtitle row (2nd)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_columns)
    subtitle_cell = ws.cell(row=2, column=1)
    subtitle_cell.value = "Detailed personnel deduction adjustments"
    subtitle_cell.font = Font(size=10, italic=True)
    subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 20

    # Use ExportUtilityService to split header lines on row 3
    export_util = ExportUtilityService()
    ws.append([export_util.split_header_to_lines(h) for h in headers])

    # Header styling
    header_fill = PatternFill(start_color="FF0070C0", end_color="FF0070C0", fill_type="solid")  # Blue fill
    header_font = Font(bold=True, color="FFFFFFFF")  # White font
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Add data rows starting at row 4
    for i, d in enumerate(deductions, start=1):
        ws.append([
            i,
            getattr(d.record_month.payroll_month, "payroll_month", ""),
            getattr(d.payroll_needing_adjustment.payroll_month, "payroll_month", ""),
            getattr(d.record_month.personnel_full_name, "first_name", ""),
            getattr(d.record_month.personnel_full_name, "father_name", ""),
            getattr(d.record_month.personnel_full_name, "last_name", ""),
            d.get_case_display(),
            d.get_component_display(),
            float(d.deduction_amount or 0),
            d.period_start.strftime("%Y-%m-%d") if d.period_start else "",
            d.period_end.strftime("%Y-%m-%d") if d.period_end else "",
            d.months_covered,
            d.created_at.strftime("%Y-%m-%d %H:%M") if d.created_at else "",
            d.updated_at.strftime("%Y-%m-%d %H:%M") if d.updated_at else "",
        ])

    # Adjust column widths
    MIN_WIDTH = 10
    MAX_WIDTH = 25
    for i, column_cells in enumerate(ws.columns, 1):
        max_length = 0
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                continue
        adjusted_width = min(max(max_length + 2, MIN_WIDTH), MAX_WIDTH)
        ws.column_dimensions[get_column_letter(i)].width = adjusted_width

    # Output to HttpResponse
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response['Content-Disposition'] = 'attachment; filename=deduction_adjustments.xlsx'
    return response



@login_required
def export_deduction_per_adjusted_month_to_excel(request):
    context = get_deduction_adjustment_context(request)
    data = context.get('deduction_per_adjusted_month', [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Deduction Per Adjusted Payroll Month"

    headers = [
        "Record Month",
        "Adjusted Payroll Month",
        "Personnel ID",
        "First Name",
        "Father Name",
        "Last Name",
        "Adjusted Deduction Per Adjusted Payroll Month"
    ]

    total_columns = len(headers)

    # Title row (1st)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "Deduction Per Adjusted Payroll Month Report"
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Subtitle row (2nd)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_columns)
    subtitle_cell = ws.cell(row=2, column=1)
    subtitle_cell.value = "Detailed personnel deduction adjustments per payroll month"
    subtitle_cell.font = Font(size=10, italic=True)
    subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 20

    # Use ExportUtilityService to split header lines
    export_util = ExportUtilityService()
    ws.append([export_util.split_header_to_lines(h) for h in headers])

    # Header style
    header_fill = PatternFill(start_color="FF0070C0", end_color="FF0070C0", fill_type="solid")  # Blue fill
    header_font = Font(bold=True, color="FFFFFFFF")  # White font
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Append data rows starting from row 4
    for item in data:
        ws.append([
            item.get("record_month__payroll_month__payroll_month", ""),
            item.get("payroll_needing_adjustment__payroll_month__payroll_month", ""),
            item.get("record_month__personnel_full_name__personnel_id", ""),
            item.get("record_month__personnel_full_name__first_name", ""),
            item.get("record_month__personnel_full_name__father_name", ""),
            item.get("record_month__personnel_full_name__last_name", ""),
            float(item.get("adjusted_deduction_per_adjusted_month", 0) or 0),
        ])

    # Auto-adjust column widths
    MIN_WIDTH = 12
    MAX_WIDTH = 40
    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_length = 0
        for cell in col_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max(max_length + 2, MIN_WIDTH), MAX_WIDTH)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # Save workbook to memory and return response
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = "attachment; filename=deduction_per_adjusted_month.xlsx"
    return response



@login_required
def export_monthly_deduction_adjustment_to_excel(request):
    context = get_deduction_adjustment_context(request)
    data = context.get('monthly_deduction_adjustment', [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Deduction Adjustment"

    headers = [
        "Payroll Month",
        "Personnel ID",
        "First Name",
        "Father Name",
        "Last Name",
        "Monthly Adjusted Deduction",
    ]

    total_columns = len(headers)

    # Title row (1st)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "Monthly Deduction Adjustment Report"
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Subtitle row (2nd)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_columns)
    subtitle_cell = ws.cell(row=2, column=1)
    subtitle_cell.value = "Details of personnel monthly deduction adjustments"
    subtitle_cell.font = Font(size=10, italic=True)
    subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 20

    # Use ExportUtilityService to split header lines
    export_util = ExportUtilityService()
    ws.append([export_util.split_header_to_lines(h) for h in headers])

    # Header style
    header_fill = PatternFill(start_color="FF0070C0", end_color="FF0070C0", fill_type="solid")  # Blue fill
    header_font = Font(bold=True, color="FFFFFFFF")  # White font
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Append data rows
    for item in data:
        ws.append([
            item.get("record_month__payroll_month__payroll_month", ""),
            item.get("record_month__personnel_full_name__personnel_id", ""),
            item.get("record_month__personnel_full_name__first_name", ""),
            item.get("record_month__personnel_full_name__father_name", ""),
            item.get("record_month__personnel_full_name__last_name", ""),
            float(item.get("monthly_adjusted_deduction", 0) or 0),
        ])

    # Adjust column widths
    MIN_WIDTH = 12
    MAX_WIDTH = 40
    for i, col_cells in enumerate(ws.columns, 1):
        max_length = 0
        for cell in col_cells:
            try:
                if cell.value:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            except Exception:
                pass
        adjusted_width = max(MIN_WIDTH, min(MAX_WIDTH, max_length + 2))
        ws.column_dimensions[get_column_letter(i)].width = adjusted_width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = "attachment; filename=monthly_deduction_adjustment.xlsx"
    return response
